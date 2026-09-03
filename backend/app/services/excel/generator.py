import io
from decimal import Decimal
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from backend.app.models.trip import Trip, TripMember
from backend.app.models.itinerary import ItineraryDay, ItineraryItem
from backend.app.models.flight import Flight
from backend.app.models.expense import Expense
from backend.app.services.expenses.calculator import expense_calculator

class ExcelExportService:
    @staticmethod
    def generate_expense_workbook(db: Session, trip_id: str) -> io.BytesIO:
        trip = db.query(Trip).filter(Trip.id == trip_id).first()
        if not trip:
            raise ValueError("Trip not found")

        analytics = expense_calculator.calculate_trip_analytics(db, trip_id)
        expenses = db.query(Expense).filter(Expense.trip_id == trip_id).order_by(Expense.expense_date.asc()).all()
        members = db.query(TripMember).filter(TripMember.trip_id == trip_id).all()
        days = db.query(ItineraryDay).filter(ItineraryDay.trip_id == trip_id).order_by(ItineraryDay.day_number.asc()).all()
        flights = db.query(Flight).filter(Flight.trip_id == trip_id).order_by(Flight.departure_time.asc()).all()

        wb = Workbook()

        # Premium Executive Color Palettes
        NAVY_FILL = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        HEADER_FILL = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        SUBHEADER_FILL = PatternFill(start_color="312E81", end_color="312E81", fill_type="solid")
        LIGHT_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        ACCENT_FILL = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
        SUCCESS_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")

        TITLE_FONT = Font(name="Segoe UI", size=13, bold=True, color="FFFFFF")
        WHITE_FONT_BOLD = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        BOLD_FONT = Font(name="Segoe UI", size=10, bold=True, color="0F172A")
        REGULAR_FONT = Font(name="Segoe UI", size=10, color="1E293B")
        MUTED_FONT = Font(name="Segoe UI", size=9, italic=True, color="64748B")

        THIN_BORDER = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1"),
        )
        DOUBLE_BOTTOM_BORDER = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="double", color="0F172A"),
        )

        def apply_banner(ws, end_col_letter: str, title_text: str):
            ws.merge_cells(f"A1:{end_col_letter}1")
            banner_cell = ws["A1"]
            banner_cell.value = title_text
            banner_cell.font = TITLE_FONT
            banner_cell.fill = NAVY_FILL
            banner_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 30

        def apply_headers(ws, headers: list, row_num: int = 3, fill = HEADER_FILL):
            for col_idx, header in enumerate(headers, 1):
                c = ws.cell(row=row_num, column=col_idx, value=header)
                c.font = WHITE_FONT_BOLD
                c.fill = fill
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = THIN_BORDER
            ws.row_dimensions[row_num].height = 22

        # ----------------------------------------------------
        # SHEET 1: TRIP OVERVIEW & EXECUTIVE SUMMARY
        # ----------------------------------------------------
        ws_overview = wb.active
        ws_overview.title = "Trip Overview"
        ws_overview.views.sheetView[0].showGridLines = True
        apply_banner(ws_overview, "E", f"ARC-NOMADE — {trip.title.upper()} (OVERVIEW)")

        # Trip Key Information Card
        info_items = [
            ("Trip Name", trip.title),
            ("Destination", f"{trip.destination} ({trip.destination_lat:.4f}, {trip.destination_lng:.4f})" if trip.destination_lat and trip.destination_lng else trip.destination),
            ("Dates", f"{trip.start_date.strftime('%b %d, %Y')} – {trip.end_date.strftime('%b %d, %Y')}"),
            ("Duration", f"{(trip.end_date - trip.start_date).days + 1} Days"),
            ("Status", trip.status),
            ("Primary Currency", trip.currency),
            ("Total Group Budget", f"{trip.currency} {float(trip.budget):,.2f}"),
            ("Total Actual Spent", f"{trip.currency} {float(analytics.total_spent):,.2f}"),
            ("Remaining Budget", f"{trip.currency} {float(trip.budget - analytics.total_spent):,.2f}"),
            ("Budget Utilization", f"{(float(analytics.total_spent) / float(trip.budget) * 100):.1f}%" if trip.budget and trip.budget > 0 else "N/A"),
        ]

        ws_overview.cell(row=3, column=1, value="KEY METRICS").font = BOLD_FONT
        ws_overview.cell(row=3, column=2, value="VALUE").font = BOLD_FONT
        for col_idx in [1, 2]:
            ws_overview.cell(row=3, column=col_idx).fill = ACCENT_FILL
            ws_overview.cell(row=3, column=col_idx).border = THIN_BORDER
        ws_overview.row_dimensions[3].height = 20

        curr_row = 4
        for label, val in info_items:
            ws_overview.cell(row=curr_row, column=1, value=label).font = BOLD_FONT
            ws_overview.cell(row=curr_row, column=1).alignment = Alignment(horizontal="left")
            ws_overview.cell(row=curr_row, column=1).border = THIN_BORDER
            ws_overview.cell(row=curr_row, column=2, value=val).font = REGULAR_FONT
            ws_overview.cell(row=curr_row, column=2).alignment = Alignment(horizontal="left")
            ws_overview.cell(row=curr_row, column=2).border = THIN_BORDER
            curr_row += 1

        # Trip Members Table
        curr_row += 2
        ws_overview.cell(row=curr_row, column=1, value="TRIP PARTICIPANTS & ROLES").font = BOLD_FONT
        curr_row += 1
        headers_m = ["Participant Name", "Email", "Trip Role", "Joined Date"]
        for col_idx, h in enumerate(headers_m, 1):
            c = ws_overview.cell(row=curr_row, column=col_idx, value=h)
            c.font = WHITE_FONT_BOLD
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = THIN_BORDER
        ws_overview.row_dimensions[curr_row].height = 20

        curr_row += 1
        for m in members:
            name = m.user.full_name if m.user else "Member"
            email = m.user.email if m.user else "—"
            ws_overview.cell(row=curr_row, column=1, value=name).alignment = Alignment(horizontal="left")
            ws_overview.cell(row=curr_row, column=2, value=email).alignment = Alignment(horizontal="left")
            ws_overview.cell(row=curr_row, column=3, value=m.role).alignment = Alignment(horizontal="center")
            ws_overview.cell(row=curr_row, column=4, value=m.joined_at.strftime("%Y-%m-%d") if m.joined_at else "—").alignment = Alignment(horizontal="center")
            for col_idx in range(1, 5):
                ws_overview.cell(row=curr_row, column=col_idx).font = REGULAR_FONT
                ws_overview.cell(row=curr_row, column=col_idx).border = THIN_BORDER
            curr_row += 1

        # ----------------------------------------------------
        # SHEET 2: ITINERARY & DAILY SCHEDULE
        # ----------------------------------------------------
        ws_itin = wb.create_sheet(title="Itinerary Schedule")
        ws_itin.views.sheetView[0].showGridLines = True
        apply_banner(ws_itin, "I", f"ARC-NOMADE — {trip.title.upper()} (ITINERARY SCHEDULE)")

        itin_headers = ["Day #", "Day Theme / Focus", "Time Window", "Category", "Activity / Stop", "Location / Neighborhood", "Estimated Cost", "Currency", "Notes"]
        apply_headers(ws_itin, itin_headers, row_num=3)
        ws_itin.freeze_panes = "A4"

        row_it = 4
        all_items_count = 0
        for day in days:
            day_items = day.items if day.items else []
            if not day_items:
                ws_itin.cell(row=row_it, column=1, value=f"Day {day.day_number}").alignment = Alignment(horizontal="center")
                ws_itin.cell(row=row_it, column=2, value=day.notes or "Free exploration day").alignment = Alignment(horizontal="left")
                ws_itin.cell(row=row_it, column=3, value="All Day").alignment = Alignment(horizontal="center")
                ws_itin.cell(row=row_it, column=4, value="RELAXATION").alignment = Alignment(horizontal="center")
                ws_itin.cell(row=row_it, column=5, value="Free Time & Exploration").alignment = Alignment(horizontal="left")
                ws_itin.cell(row=row_it, column=6, value=trip.destination).alignment = Alignment(horizontal="left")
                ws_itin.cell(row=row_it, column=7, value=0.0).number_format = "#,##0.00"
                ws_itin.cell(row=row_it, column=8, value=trip.currency).alignment = Alignment(horizontal="center")
                ws_itin.cell(row=row_it, column=9, value="No specific activities scheduled yet.").alignment = Alignment(horizontal="left")
                for col_idx in range(1, 10):
                    ws_itin.cell(row=row_it, column=col_idx).font = REGULAR_FONT
                    ws_itin.cell(row=row_it, column=col_idx).border = THIN_BORDER
                row_it += 1
            else:
                for item in day_items:
                    all_items_count += 1
                    time_win = f"{item.start_time or ''} - {item.end_time or ''}".strip(" -")
                    ws_itin.cell(row=row_it, column=1, value=f"Day {day.day_number}").alignment = Alignment(horizontal="center")
                    ws_itin.cell(row=row_it, column=2, value=day.notes or "").alignment = Alignment(horizontal="left")
                    ws_itin.cell(row=row_it, column=3, value=time_win or "Flexible").alignment = Alignment(horizontal="center")
                    ws_itin.cell(row=row_it, column=4, value=item.category or "SIGHTSEEING").alignment = Alignment(horizontal="center")
                    ws_itin.cell(row=row_it, column=5, value=item.title).alignment = Alignment(horizontal="left")
                    ws_itin.cell(row=row_it, column=6, value=item.location_name or item.address or trip.destination).alignment = Alignment(horizontal="left")
                    cost_cell = ws_itin.cell(row=row_it, column=7, value=float(item.estimated_cost or 0.0))
                    cost_cell.number_format = "#,##0.00"
                    cost_cell.alignment = Alignment(horizontal="right")
                    ws_itin.cell(row=row_it, column=8, value=item.currency or trip.currency).alignment = Alignment(horizontal="center")
                    ws_itin.cell(row=row_it, column=9, value=item.notes or item.description or "").alignment = Alignment(horizontal="left")
                    for col_idx in range(1, 10):
                        ws_itin.cell(row=row_it, column=col_idx).font = REGULAR_FONT
                        ws_itin.cell(row=row_it, column=col_idx).border = THIN_BORDER
                        if row_it % 2 == 0:
                            ws_itin.cell(row=row_it, column=col_idx).fill = LIGHT_FILL
                    row_it += 1

        if row_it > 4:
            ws_itin.auto_filter.ref = f"A3:I{row_it - 1}"

        # ----------------------------------------------------
        # SHEET 3: FLIGHTS & LOGISTICS
        # ----------------------------------------------------
        ws_flights = wb.create_sheet(title="Flights & Logistics")
        ws_flights.views.sheetView[0].showGridLines = True
        apply_banner(ws_flights, "I", f"ARC-NOMADE — {trip.title.upper()} (FLIGHTS & LOGISTICS)")

        flight_headers = ["Airline", "Flight #", "Departure Airport", "Arrival Airport", "Departure Time", "Arrival Time", "Terminal / Gate", "Seat", "Status"]
        apply_headers(ws_flights, flight_headers, row_num=3)
        ws_flights.freeze_panes = "A4"

        row_fl = 4
        if not flights:
            ws_flights.cell(row=row_fl, column=1, value="No flights logged for this journey yet.").font = MUTED_FONT
            row_fl += 1
        else:
            for f in flights:
                term_gate = f"{f.terminal or ''} / {f.gate or ''}".strip(" /")
                ws_flights.cell(row=row_fl, column=1, value=f.airline).alignment = Alignment(horizontal="left")
                ws_flights.cell(row=row_fl, column=2, value=f.flight_number).alignment = Alignment(horizontal="center")
                ws_flights.cell(row=row_fl, column=3, value=f.departure_airport).alignment = Alignment(horizontal="center")
                ws_flights.cell(row=row_fl, column=4, value=f.arrival_airport).alignment = Alignment(horizontal="center")
                ws_flights.cell(row=row_fl, column=5, value=f.departure_time.strftime("%Y-%m-%d %H:%M") if f.departure_time else "").alignment = Alignment(horizontal="center")
                ws_flights.cell(row=row_fl, column=6, value=f.arrival_time.strftime("%Y-%m-%d %H:%M") if f.arrival_time else "").alignment = Alignment(horizontal="center")
                ws_flights.cell(row=row_fl, column=7, value=term_gate or "—").alignment = Alignment(horizontal="center")
                ws_flights.cell(row=row_fl, column=8, value=f.seat or "—").alignment = Alignment(horizontal="center")
                ws_flights.cell(row=row_fl, column=9, value=f.status).alignment = Alignment(horizontal="center")
                for col_idx in range(1, 10):
                    ws_flights.cell(row=row_fl, column=col_idx).font = REGULAR_FONT
                    ws_flights.cell(row=row_fl, column=col_idx).border = THIN_BORDER
                    if row_fl % 2 == 0:
                        ws_flights.cell(row=row_fl, column=col_idx).fill = LIGHT_FILL
                row_fl += 1

        if row_fl > 4:
            ws_flights.auto_filter.ref = f"A3:I{row_fl - 1}"

        # ----------------------------------------------------
        # SHEET 4: EXPENSE LEDGER
        # ----------------------------------------------------
        ws_exp = wb.create_sheet(title="Expense Ledger")
        ws_exp.views.sheetView[0].showGridLines = True
        apply_banner(ws_exp, "G", f"ARC-NOMADE — {trip.title.upper()} (EXPENSE LEDGER)")

        headers_exp = ["Date", "Category", "Description", "Paid By", "Amount", "Currency", "Split Type"]
        apply_headers(ws_exp, headers_exp, row_num=3)
        ws_exp.freeze_panes = "A4"

        row_idx = 4
        for exp in expenses:
            ws_exp.cell(row=row_idx, column=1, value=exp.expense_date.strftime("%Y-%m-%d")).alignment = Alignment(horizontal="center")
            ws_exp.cell(row=row_idx, column=2, value=exp.category).alignment = Alignment(horizontal="center")
            ws_exp.cell(row=row_idx, column=3, value=exp.description).alignment = Alignment(horizontal="left")
            ws_exp.cell(row=row_idx, column=4, value=exp.payer.full_name if exp.payer else "Member").alignment = Alignment(horizontal="left")
            
            amt_cell = ws_exp.cell(row=row_idx, column=5, value=float(exp.amount))
            amt_cell.number_format = "#,##0.00"
            amt_cell.alignment = Alignment(horizontal="right")

            ws_exp.cell(row=row_idx, column=6, value=exp.currency).alignment = Alignment(horizontal="center")
            ws_exp.cell(row=row_idx, column=7, value=exp.split_type).alignment = Alignment(horizontal="center")

            for col_num in range(1, 8):
                ws_exp.cell(row=row_idx, column=col_num).font = REGULAR_FONT
                ws_exp.cell(row=row_idx, column=col_num).border = THIN_BORDER
                if row_idx % 2 == 0:
                    ws_exp.cell(row=row_idx, column=col_num).fill = LIGHT_FILL
            row_idx += 1

        # Dynamic Formula for Total Spent
        ws_exp.cell(row=row_idx, column=1, value="TOTAL SPENT").font = BOLD_FONT
        ws_exp.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left")
        
        total_cell = ws_exp.cell(row=row_idx, column=5)
        if row_idx > 4:
            total_cell.value = f"=SUM(E4:E{row_idx - 1})"
            ws_exp.auto_filter.ref = f"A3:G{row_idx - 1}"
        else:
            total_cell.value = 0.0

        total_cell.font = BOLD_FONT
        total_cell.number_format = "#,##0.00"
        total_cell.alignment = Alignment(horizontal="right")
        ws_exp.cell(row=row_idx, column=6, value=trip.currency).font = BOLD_FONT
        ws_exp.cell(row=row_idx, column=6).alignment = Alignment(horizontal="center")

        for col_num in range(1, 8):
            ws_exp.cell(row=row_idx, column=col_num).border = DOUBLE_BOTTOM_BORDER
            ws_exp.cell(row=row_idx, column=col_num).fill = ACCENT_FILL

        # ----------------------------------------------------
        # SHEET 5: MEMBER BALANCES & SETTLEMENTS
        # ----------------------------------------------------
        ws_bal = wb.create_sheet(title="Balances & Settlements")
        ws_bal.views.sheetView[0].showGridLines = True
        apply_banner(ws_bal, "E", "MEMBER BALANCE SUMMARY & OPTIMIZED SETTLEMENT PLAN")

        headers_bal = ["Member Name", "Total Paid", "Total Share", "Net Balance", "Settlement Status"]
        apply_headers(ws_bal, headers_bal, row_num=3)
        ws_bal.freeze_panes = "A4"

        r2 = 4
        for mb in analytics.spending_by_member:
            ws_bal.cell(row=r2, column=1, value=mb.user_name).alignment = Alignment(horizontal="left")
            
            p_cell = ws_bal.cell(row=r2, column=2, value=float(mb.total_paid))
            p_cell.number_format = "#,##0.00"
            p_cell.alignment = Alignment(horizontal="right")

            s_cell = ws_bal.cell(row=r2, column=3, value=float(mb.total_share))
            s_cell.number_format = "#,##0.00"
            s_cell.alignment = Alignment(horizontal="right")

            net_cell = ws_bal.cell(row=r2, column=4, value=float(mb.net_balance))
            net_cell.number_format = "#,##0.00"
            net_cell.alignment = Alignment(horizontal="right")
            net_cell.font = BOLD_FONT

            status_str = "Settled" if abs(mb.net_balance) < Decimal("0.01") else (
                f"Gets back {trip.currency} {mb.net_balance:,.2f}" if mb.net_balance > 0 else f"Owes {trip.currency} {-mb.net_balance:,.2f}"
            )
            ws_bal.cell(row=r2, column=5, value=status_str).alignment = Alignment(horizontal="left")

            for col_num in range(1, 6):
                ws_bal.cell(row=r2, column=col_num).font = REGULAR_FONT if col_num != 4 else BOLD_FONT
                ws_bal.cell(row=r2, column=col_num).border = THIN_BORDER
                if r2 % 2 == 0:
                    ws_bal.cell(row=r2, column=col_num).fill = LIGHT_FILL
            r2 += 1

        # Suggested Settlements Sub-table
        r2 += 2
        ws_bal.merge_cells(f"A{r2}:D{r2}")
        settle_hdr = ws_bal.cell(row=r2, column=1, value="OPTIMIZED SETTLEMENT TRANSACTIONS (MINIMUM CASH FLOW)")
        settle_hdr.font = WHITE_FONT_BOLD
        settle_hdr.fill = SUBHEADER_FILL
        settle_hdr.alignment = Alignment(horizontal="center", vertical="center")
        ws_bal.row_dimensions[r2].height = 20

        r2 += 1
        headers_set = ["Payer (Who Pays)", "Receiver (Who Gets Paid)", "Amount", "Currency"]
        for col_num, header in enumerate(headers_set, 1):
            c = ws_bal.cell(row=r2, column=col_num)
            c.value = header
            c.font = WHITE_FONT_BOLD
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = THIN_BORDER
        ws_bal.row_dimensions[r2].height = 20

        r2 += 1
        if not analytics.suggested_settlements:
            ws_bal.cell(row=r2, column=1, value="All trip expenses are currently fully settled!").font = MUTED_FONT
        else:
            for s in analytics.suggested_settlements:
                ws_bal.cell(row=r2, column=1, value=s.payer_name).alignment = Alignment(horizontal="left")
                ws_bal.cell(row=r2, column=2, value=s.receiver_name).alignment = Alignment(horizontal="left")
                amt_c = ws_bal.cell(row=r2, column=3, value=float(s.amount))
                amt_c.number_format = "#,##0.00"
                amt_c.alignment = Alignment(horizontal="right")
                amt_c.font = BOLD_FONT
                ws_bal.cell(row=r2, column=4, value=s.currency).alignment = Alignment(horizontal="center")
                for col_num in range(1, 5):
                    ws_bal.cell(row=r2, column=col_num).border = THIN_BORDER
                    ws_bal.cell(row=r2, column=col_num).font = REGULAR_FONT if col_num != 3 else BOLD_FONT
                r2 += 1

        # ----------------------------------------------------
        # SHEET 6: CATEGORY ANALYTICS
        # ----------------------------------------------------
        ws_cat = wb.create_sheet(title="Category Breakdown")
        ws_cat.views.sheetView[0].showGridLines = True
        apply_banner(ws_cat, "D", "SPENDING BY CATEGORY & BUDGET UTILIZATION")

        headers_cat = ["Category", "Total Spent", "Percentage (%)", "Trip Currency"]
        apply_headers(ws_cat, headers_cat, row_num=3)
        ws_cat.freeze_panes = "A4"

        r3 = 4
        for cat in analytics.spending_by_category:
            ws_cat.cell(row=r3, column=1, value=cat.category).alignment = Alignment(horizontal="left")
            c_amt = ws_cat.cell(row=r3, column=2, value=float(cat.amount))
            c_amt.number_format = "#,##0.00"
            c_amt.alignment = Alignment(horizontal="right")
            
            c_pct = ws_cat.cell(row=r3, column=3, value=f"{cat.percentage:.1f}%")
            c_pct.alignment = Alignment(horizontal="right")
            ws_cat.cell(row=r3, column=4, value=trip.currency).alignment = Alignment(horizontal="center")

            for col_num in range(1, 5):
                ws_cat.cell(row=r3, column=col_num).font = REGULAR_FONT
                ws_cat.cell(row=r3, column=col_num).border = THIN_BORDER
                if r3 % 2 == 0:
                    ws_cat.cell(row=r3, column=col_num).fill = LIGHT_FILL
            r3 += 1

        # ----------------------------------------------------
        # PROPORTIONAL COLUMN WIDTH CALCULATION (Fix merged cell width bug)
        # ----------------------------------------------------
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    # Skip merged header and title rows so they don't stretch Column A
                    if cell.row in [1, 2]:
                        continue
                    if cell.value is not None:
                        val_str = str(cell.value)
                        # Skip section banner rows that span across columns
                        if len(val_str) > 40 and cell.column == 1:
                            continue
                        if not val_str.startswith("="):
                            max_len = max(max_len, len(val_str))
                sheet.column_dimensions[col_letter].width = max(min(max_len + 5, 42), 14)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

excel_export_service = ExcelExportService()
